# د رقم الجلوس د نقشې د جوړولو سیستم
# ──────────────────────────────────────────────────────────────────────────────
#  د زده کوونکو د لورډولو سیستم:
#    • د امتحان او صنف انتخاب (له vw_ClassesInExams څخه)
#    • زده کوونکي له vw_AdmittedStudentsInExams څخه
#  ټول نور ځانګړتیاوې بدلون نه لري.
# ──────────────────────────────────────────────────────────────────────────────

import os, sys, random, textwrap
import threading
import time
from pathlib import Path
from ortools.sat.python import cp_model

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from PyQt5.QtCore import Qt, QRect, QPoint, QEvent, QSize, pyqtSignal
from PyQt5.QtGui import QColor, QFont, QPainter, QPen, QCursor, QKeySequence, QMouseEvent
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QLabel,
    QPushButton,
    QVBoxLayout,
    QHBoxLayout,
    QFrame,
    QGridLayout,
    QComboBox,
    QColorDialog,
    QMessageBox,
    QInputDialog,
    QLineEdit,
    QMenu,
    QScrollArea,
    QShortcut,
    QSizePolicy,
    QFileDialog,
    QCheckBox,
    QDialog,
    QRubberBand,
    QTableWidget,
    QTableWidgetItem,
    QProgressDialog,
    QTableWidgetItem,
    QGroupBox,
    QTabWidget,
    QAction,
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from nazim.db_singleton import DBManager
from nazim.modules.rollnumber.SeatingMapReportDialog import SeatingMapReportDialog

# ───────────────────────── constants ───────────────────────────────────────────
UNDO_LIMIT = 30
VIEW_CLASSES = "vw_ClassesInExams"  # should expose exam_id, class_name
VIEW_STUDS = (
    "vw_AdmittedStudentsInExams"  # exam_id, student_id, roll_number, class_name, name, father_name
)

# 8-directional neighbour offsets (orthogonal + diagonal) — a diagonal neighbour
# can still see another student's paper, so it counts as "adjacent".
ADJ_DIRECTIONS = ((-1, -1), (-1, 0), (-1, 1), (0, -1), (0, 1), (1, -1), (1, 0), (1, 1))


# ───────────────────────── seat widget ────────────────────────────────────────
class CellWidget(QFrame):
    changed = pyqtSignal()
    right_click = pyqtSignal(object)

    def __init__(self, seat_no: int, colour_fn, parent_app):
        super().__init__()
        self.number = seat_no
        self.get_colour = colour_fn
        self.parent_app = parent_app
        self.student_data = None
        self.locked = False
        self.selected = False
        self.search_highlight = False  # New flag for search highlighting

        self.setFrameStyle(QFrame.Panel | QFrame.Sunken)
        self.setLineWidth(1)
        self.setMinimumSize(40, 12)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.setMouseTracking(True)
        self.setToolTip("")

    def paintEvent(self, _):
        p = QPainter(self)
        
        # Set light border for all cells
        p.setPen(QPen(QColor("#CCCCCC"), 1))
        
        if self.student_data:
            col = self.get_colour(self.student_data["Class"])
            p.fillRect(self.rect(), QColor(f"#{col}"))
        elif self.locked:
            p.fillRect(self.rect(), QColor("#AAAAAA"))
        else:
            p.fillRect(self.rect(), QColor("#ECF0F1"))
        
        # Draw border - use red border for search highlight
        if self.search_highlight:
            p.setPen(QPen(QColor("red"), 3))
        else:
            p.setPen(QPen(QColor("#CCCCCC"), 1))
        p.drawRect(self.rect())
        
        # Draw text
        p.setPen(Qt.black)
        p.drawText(self.rect(), Qt.AlignCenter | Qt.TextWordWrap, str(self.number))
        
        if self.selected:
            p.setPen(QPen(Qt.blue, 2))
            p.drawRect(1, 1, self.width() - 2, self.height() - 2)

    def enterEvent(self, _):
        if self.student_data:
            d = self.student_data
            tip = (
                f"ID: {d['ID']}\n"
                f"Name: {d['Name']}\n"
                f"Father: {d['Father']}\n"
                f"Class: {d['Class']}\n"
                f"Seat: {self.number}"
            )
            self.setToolTip(tip)
        else:
            self.setToolTip("")
        super().enterEvent(_)

    def mouseDoubleClickEvent(self, e):
        if e.button() == Qt.LeftButton and self.student_data:
            self.parent_app.start_move(self)
        super().mouseDoubleClickEvent(e)

    def mousePressEvent(self, e):
        if e.button() == Qt.LeftButton:
            if self.parent_app.moving_cell:
                self.parent_app.try_swap(self)
            else:
                self.selected = not self.selected
                self.update()
        super().mousePressEvent(e)

    def contextMenuEvent(self, _):
        self.right_click.emit(self)

    def clear(self):
        self.student_data = None
        self.changed.emit()
        self.update()

    def toggle_lock(self):
        self.locked = not self.locked
        self.changed.emit()
        self.update()


# ─────────────────── grid wrapper for drag‐to‐select ──────────────────────────
class GridWidget(QWidget):
    def __init__(self, hall, layout, parent_app):
        super().__init__()
        self.hall = hall
        self.parent_app = parent_app
        self.setLayout(layout)
        self._dragging = False
        self._start = QPoint()
        self._end = QPoint()
        self._mode = "select"
        self.rubberBand = QRubberBand(QRubberBand.Rectangle, self)
        for row in self.hall:
            for cw in row:
                cw.installEventFilter(self)

    def eventFilter(self, obj, event):
        if event.type() == QEvent.MouseButtonPress and isinstance(event, QMouseEvent):
            if event.button() == Qt.LeftButton:
                return self._handlePress(obj, event)
        if event.type() == QEvent.MouseMove and isinstance(event, QMouseEvent):
            if event.buttons() & Qt.LeftButton:
                return self._handleMove(obj, event)
        if event.type() == QEvent.MouseButtonRelease and isinstance(event, QMouseEvent):
            if event.button() == Qt.LeftButton:
                return self._handleRelease(obj, event)
        return super().eventFilter(obj, event)

    def _handlePress(self, obj, event):
        # record where the click happened
        self._start = obj.mapToParent(event.pos())
        self._end = self._start

        # check if click was on any cell
        hit_cell = False
        for row in self.hall:
            for cw in row:
                rect = QRect(cw.mapToParent(QPoint(0, 0)), cw.size())
                if rect.contains(self._start):
                    hit_cell = True
                    # toggle or set select mode based on current state
                    self._mode = "deselect" if cw.selected else "select"
                    break
            if hit_cell:
                break

        if not hit_cell:
            # clear all selections if clicked outside
            for row in self.hall:
                for cw in row:
                    if cw.selected:
                        cw.selected = False
                        cw.update()
            return True

        # otherwise start a new rubber‐band drag
        self._dragging = True
        self.rubberBand.setGeometry(QRect(self._start, QSize()))
        self.rubberBand.show()
        self._update_selection()
        return True

    def _handleMove(self, obj, event):
        self._end = obj.mapToParent(event.pos())
        self.rubberBand.setGeometry(QRect(self._start, self._end).normalized())
        self._update_selection()
        return True

    def _handleRelease(self, obj, event):
        self._end = obj.mapToParent(event.pos())
        self._update_selection()
        self._dragging = False
        self.rubberBand.hide()
        return True

    def _update_selection(self):
        rect = QRect(self._start, self._end).normalized()
        for row in self.hall:
            for cw in row:
                top_left = cw.mapToParent(QPoint(0, 0))
                if rect.intersects(QRect(top_left, cw.size())):
                    cw.selected = self._mode == "select"
                    cw.update()


# ───────────────────────── main window ────────────────────────────────────────
class RollNumberMappingApp(QMainWindow):
    seatAssignmentResult = pyqtSignal(list, list, bool)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("د رقم الجلوس د نقشې د جوړولو سیستم")
        self.setGeometry(100, 100, 1200, 800)
        self.setMinimumSize(800, 600)

        # Define fonts and colors (EXACTLY as in general registration)
        self.label_font = QFont("Bahij Nassim", 14)
        self.label_font.setBold(True)
        self.input_font = QFont("Bahij Nassim", 12)
        self.title_font = QFont("Bahij Titr", 24)
        
        self.label_color = "color: #0b0b56;"  # Dark blue color for labels
        self.button_style = "background-color: #0b0b56; color: white;"  # Matching buttons
        self.title_color = "color: #0b0b56;"  # Dark blue for titles
        
        self.message_box_style = """
            QMessageBox {
                font-family: 'Bahij Nassim';
                font-size: 12pt;
            }
            QMessageBox QLabel {
                color: #0b0b56;
            }
            QMessageBox QPushButton {
                background-color: #0b0b56;
                color: white;
                font-size: 12pt;
                min-width: 80px;
            }
        """

        # DB map id
        self.current_map_id = None
        # initial grid
        self.rows, self.cols = 40, 20
        # state
        self.student_data = []  # list of dicts
        self.class_colors = {}
        self.hall_layout = []
        self.moving_cell = None
        self.copied_data = None
        self.undo_stack = []
        self.redo_stack = []

        self.build_ui()

    # ───────────────────── build UI ─────────────────────────
    def build_ui(self):
        # Main widget and layout
        main_widget = QWidget(self)
        main_layout = QVBoxLayout(main_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        # Title with separator (EXACTLY as in general registration)
        title_label = QLabel("د رقم الجلوس د نقشې د جوړولو سیستم")
        title_label.setFont(self.title_font)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet(self.title_color)
        main_layout.addWidget(title_label)

        # Add a light border (separator)
        separator = QFrame(self)
        separator.setFrameShape(QFrame.HLine)
        separator.setFrameShadow(QFrame.Sunken)
        separator.setStyleSheet("color: #0b0b56;")
        separator.setLineWidth(1)
        main_layout.addWidget(separator)

        # Combined controls in one line: Search, Exam/Class, Actions
        controls_row = QHBoxLayout()
        controls_row.setContentsMargins(5, 5, 5, 5)
        controls_row.setSpacing(10)
        
        # Search group box
        search_group = QGroupBox("د زده کوونکو لټون")
        search_group.setFont(self.label_font)
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(5, 5, 5, 5)
        search_layout.setSpacing(5)
        
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("د ID یا نوم له مخی لټون...")
        self.search_box.setFont(self.input_font)
        self.search_box.setAlignment(Qt.AlignCenter)
        self.search_box.returnPressed.connect(self.perform_search)
        search_layout.addWidget(self.search_box)
        
        # Search button
        search_btn = QPushButton("لټون")
        search_btn.setFont(self.input_font)
        search_btn.setStyleSheet(self.button_style)
        search_btn.clicked.connect(self.perform_search)
        search_layout.addWidget(search_btn)
        
        search_group.setLayout(search_layout)
        controls_row.addWidget(search_group)
        
        # Exam and Class selection group box
        ctrl_group = QGroupBox("د امتحان او صنف انتخاب")
        ctrl_group.setFont(self.label_font)
        ctrl = QHBoxLayout()
        ctrl.setContentsMargins(5, 5, 5, 5)
        ctrl.setSpacing(5)
        
        # Exam selection
        exam_label = QLabel("امتحان:")
        exam_label.setFont(self.label_font)
        exam_label.setStyleSheet(self.label_color)
        ctrl.addWidget(exam_label)
        
        self.cb_exam = QComboBox()
        self.cb_exam.addItem("امتحان انتخاب کړئ", None)
        self.cb_exam.setFont(self.input_font)
        self.load_exams()
        self.cb_exam.currentIndexChanged.connect(self.on_exam_changed)
        ctrl.addWidget(self.cb_exam)
        
        # Class selection
        self.btn_select_classes = QPushButton("صنفونه انتخاب کړئ")
        self.btn_select_classes.setFont(self.input_font)
        self.btn_select_classes.setStyleSheet(self.button_style)
        self.btn_select_classes.clicked.connect(self.open_select_classes_dialog)
        ctrl.addWidget(self.btn_select_classes)
        self.selected_class_exam_ids: list[int] = []
        
        # Color selection
        self.color_btn = QPushButton("رنګونه انتخاب کړئ")
        self.color_btn.setFont(self.input_font)
        self.color_btn.setStyleSheet(self.button_style)
        self.color_btn.clicked.connect(self.set_class_colors)
        ctrl.addWidget(self.color_btn)
        
        ctrl_group.setLayout(ctrl)
        controls_row.addWidget(ctrl_group)
        
        # Actions group box
        actions_group = QGroupBox("د نقشې عملیات")
        actions_group.setFont(self.label_font)
        actions_layout = QHBoxLayout()
        actions_layout.setContentsMargins(5, 5, 5, 5)
        actions_layout.setSpacing(5)
        
        actions_layout.addWidget(self._mk_btn("نوی نقشه", self.new_map))
        actions_layout.addWidget(self._mk_btn("د رقم الجلوس نقشه جوړه کړئ", self.shuffle_and_assign_roll_numbers))
        actions_layout.addWidget(self._mk_btn("لغوه کړئ", self.undo))
        actions_layout.addWidget(self._mk_btn("بیا کړئ", self.redo))
        
        actions_group.setLayout(actions_layout)
        controls_row.addWidget(actions_group)
        
        controls_row.addStretch()
        main_layout.addLayout(controls_row)

        # Legend
        legend_group = QGroupBox("د صنفونو رنګونه")
        legend_group.setFont(self.label_font)
        self.legend_layout = QGridLayout()
        self.legend_layout.setContentsMargins(5, 5, 5, 5)
        self.legend_layout.setSpacing(5)
        legend_group.setLayout(self.legend_layout)
        main_layout.addWidget(legend_group)

        # Scrollable grid
        self.grid_layout = QGridLayout()
        self.grid_layout.setSpacing(5)
        self.rebuild_grid()
        # <<< changed: keep a reference for later rebuilds >>>
        self.grid_widget = GridWidget(self.hall_layout, self.grid_layout, self)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.grid_widget)
        scroll.setMinimumSize(1000, 600)
        scroll.setMaximumHeight(700)
        # Add horizontal scroll bar
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # Create the main content area with scroll area and right panel
        center_row = QHBoxLayout()
        center_row.setContentsMargins(0, 0, 0, 0)
        center_row.setSpacing(10)
        
        # Left side: Scroll area with grid (takes most space)
        center_row.addWidget(scroll, 4)  # Takes 4/5 of the space
        
        # Right side: Actions panel with group boxes
        right_panel = QVBoxLayout()
        right_panel.setContentsMargins(0, 0, 0, 0)
        right_panel.setSpacing(10)
        
        # Arrange group box
        arrange_group = QGroupBox("تنظیم")
        arrange_group.setFont(self.label_font)
        arrange_layout = QVBoxLayout(arrange_group)
        arrange_layout.setContentsMargins(5, 5, 5, 5)
        arrange_layout.setSpacing(5)
        
        arrange_layout.addWidget(self._mk_btn("د جدول اندازه بدل کړئ", self.change_grid_size))
        arrange_layout.addWidget(self._mk_btn("صنفونه انتخاب کړئ", self.open_select_classes_dialog))
        arrange_layout.addWidget(self._mk_btn("رنګونه انتخاب کړئ", self.set_class_colors))
        arrange_layout.addWidget(self._mk_btn("ټول لغوه کړئ", self.deselect_all))
        
        right_panel.addWidget(arrange_group)
        
        # Export group box
        export_group = QGroupBox("صادرول")
        export_group.setFont(self.label_font)
        export_layout = QVBoxLayout(export_group)
        export_layout.setContentsMargins(5, 5, 5, 5)
        export_layout.setSpacing(5)
        
        export_layout.addWidget(self._mk_btn("HTML راپور وګورئ", self.generate_html_report))
        export_layout.addWidget(self._mk_btn("Excel صادر کړئ", self.generate_excel))
        
        right_panel.addWidget(export_group)
        
        # Database group box
        db_group = QGroupBox("ډیټابیس")
        db_group.setFont(self.label_font)
        db_layout = QVBoxLayout(db_group)
        db_layout.setContentsMargins(5, 5, 5, 5)
        db_layout.setSpacing(5)
        
        db_layout.addWidget(self._mk_btn("نقشه خوندي کړئ", self.on_save_map))
        db_layout.addWidget(self._mk_btn("د خوندي شویو نقشو مدیریت", self.open_manage_maps))
        
        right_panel.addWidget(db_group)
        
        # Add stretch to push group boxes to top
        right_panel.addStretch()
        
        center_row.addLayout(right_panel, 1)  # Takes 1/5 of the space
        main_layout.addLayout(center_row)

        # Connect signals
        self.seatAssignmentResult.connect(self._on_assignment_done)

        # Set the main widget
        self.setCentralWidget(main_widget)
        
        # Shortcuts
        QShortcut(QKeySequence("Ctrl+Z"), self, activated=self.undo)
        QShortcut(QKeySequence("Ctrl+Y"), self, activated=self.redo)
        QShortcut(QKeySequence("Ctrl+S"), self, activated=self.on_save_map)

    # ───────────────────── load exams/classes ─────────────────────
    def load_exams(self):
        conn = DBManager.get_connection()
        cur = conn.cursor()
        cur.execute(f"SELECT DISTINCT exam_id FROM {VIEW_CLASSES} ORDER BY exam_id")
        for (eid,) in cur.fetchall():
            cur2 = conn.cursor()
            cur2.execute("SELECT exam_name FROM Exams WHERE id=?", eid)
            en = cur2.fetchone()
            self.cb_exam.addItem(en[0] if en else str(eid), eid)
        conn.close()

    def on_exam_changed(self, index=None):
        # Clear out any previous class selections when the user picks a new exam
        self.selected_class_exam_ids = []
        self.btn_select_classes.setText("صنفونه انتخاب کړئ")
        # Also clear out any loaded students if you want:
        self.student_data = []

        # ───────────────────── rebuild grid ──────────────────────

    def rebuild_grid(self):
        # 1) remove old widgets
        for i in reversed(range(self.grid_layout.count())):
            w = self.grid_layout.itemAt(i).widget()
            if w:
                w.setParent(None)

        # 2) rebuild layout array
        self.hall_layout = []
        seat = 1
        for r in range(self.rows):
            row = []
            for c in range(self.cols):
                cw = CellWidget(seat, self.get_class_color, self)
                cw.changed.connect(self.on_seat_change)
                cw.right_click.connect(self.show_cell_menu)
                self.grid_layout.addWidget(cw, r, c)
                row.append(cw)
                seat += 1
            self.hall_layout.append(row)

        # 3) stretch columns/rows
        for c in range(self.cols):
            self.grid_layout.setColumnStretch(c, 1)
        for r in range(self.rows):
            self.grid_layout.setRowStretch(r, 1)

        # 4) re‐attach the eventFilter to every new cell
        if hasattr(self, "grid_widget"):
            self.grid_widget.hall = self.hall_layout
            for row in self.hall_layout:
                for cw in row:
                    cw.installEventFilter(self.grid_widget)

    # ───────────────────── shuffle & assign ─────────────────────

    def shuffle_and_assign_roll_numbers(self):
        if not self.student_data:
            self.show_message_box("هیڅ معلومات نشته", "مهرباني وکړئ لومړی زده کوونکي له ډیټابیس څخه وګرځوئ.", QMessageBox.Warning)
            return

        # Get only unlocked and selected seats
        available_seats = [cw for row in self.hall_layout for cw in row if cw.selected and not cw.locked]
        m, n = len(self.student_data), len(available_seats)

        if n < m:
            # Calculate how many more cells are needed
            needed_cells = m - n
            total_selected = len([cw for row in self.hall_layout for cw in row if cw.selected])
            locked_selected = len([cw for row in self.hall_layout for cw in row if cw.selected and cw.locked])

            message = f"کافي څوکۍ نشته.\n\n"
            message += f"• د زده کوونکو شمیر: {m}\n"
            message += f"• د انتخاب شویو څوکیو شمیر: {total_selected}\n"
            message += f"• د قفل شویو څوکیو شمیر: {locked_selected}\n"
            message += f"• د موجودو څوکیو شمیر: {n}\n"
            message += f"• د اړینو څوکیو شمیر: {needed_cells}"

            self.show_message_box("کافي څوکۍ نشته", message, QMessageBox.Warning)
            return

        # ── Build seat position map in a single O(rows*cols) pass ──
        seat_index = {cw: i for i, cw in enumerate(available_seats)}
        pos = {}
        for r, row in enumerate(self.hall_layout):
            for c, cw in enumerate(row):
                if cw in seat_index:
                    pos[cw] = (r, c)

        # ── Precompute 8-directional adjacency pairs among available seats ──
        neighbor_pairs = set()
        for cw, i in seat_index.items():
            r, c = pos[cw]
            for dr, dc in ADJ_DIRECTIONS:
                nr, nc = r + dr, c + dc
                if 0 <= nr < self.rows and 0 <= nc < self.cols:
                    cw2 = self.hall_layout[nr][nc]
                    j = seat_index.get(cw2)
                    if j is not None and i < j:
                        neighbor_pairs.add((i, j))
        neighbor_pairs = list(neighbor_pairs)

        # Copy data
        students = list(self.student_data)
        seat_numbers = [cw.number for cw in available_seats]

        # Solver time budget scales with hall size but stays bounded.
        time_budget = max(5.0, min(30.0, n / 40.0))

        # Show progress dialog (indeterminate) with a cancellation bridge.
        progress = QProgressDialog(
            "د څوکیو د بهینه کولو (د یو شان صنفونو د نږدې والي له منځه وړل)...", "لغوه کړئ", 0, 0, self
        )
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        cancel_event = threading.Event()
        progress.canceled.connect(cancel_event.set)
        progress.show()

        def worker():
            try:
                best, best_cost = self._solve_seating(
                    m, n, students, neighbor_pairs, seat_numbers, cancel_event, time_budget
                )
            except Exception:
                # Any solver failure → robust heuristic fallback.
                best, best_cost = self._greedy_fallback(
                    m, n, students, neighbor_pairs, cancel_event, time_budget
                )
            fallback = best_cost > 0
            self.seatAssignmentResult.emit(best, students, fallback)

        threading.Thread(target=worker, daemon=True).start()

    # ───────────────────── CP-SAT exact solver ─────────────────────
    def _solve_seating(self, m, n, students, neighbor_pairs, seat_numbers, cancel_event, time_budget):
        """
        Assign students to seats while minimising same-class adjacency, using
        OR-Tools CP-SAT. The model works at the *class* level (students within a
        class are interchangeable for adjacency), which keeps it small and fast
        even for 1000+ seat halls.

        Returns (assignment, cost) where assignment[k] = seat index for student k
        and cost = number of remaining same-class adjacent pairs (0 when perfect).
        """
        from ortools.sat.python import cp_model

        # Group students by class (values are indices into `students`).
        classes = {}
        for k, s in enumerate(students):
            classes.setdefault(s["Class"], []).append(k)
        class_names = list(classes.keys())
        counts = {c: len(ks) for c, ks in classes.items()}

        model = cp_model.CpModel()

        # y[j, c] == 1  ⇔  seat j is occupied by a student of class c.
        y = {(j, c): model.NewBoolVar(f"y_{j}_{c}") for j in range(n) for c in class_names}

        # Each seat holds at most one class (empty seats allowed when n > m).
        for j in range(n):
            model.Add(sum(y[(j, c)] for c in class_names) <= 1)

        # Each class occupies exactly as many seats as it has students.
        for c in class_names:
            model.Add(sum(y[(j, c)] for j in range(n)) == counts[c])

        # One conflict var per adjacent seat pair: forced to 1 iff both seats
        # end up holding the same class.
        conflicts = []
        for (a, b) in neighbor_pairs:
            cf = model.NewBoolVar(f"cf_{a}_{b}")
            for c in class_names:
                model.Add(cf >= y[(a, c)] + y[(b, c)] - 1)
            conflicts.append(cf)

        model.Minimize(sum(conflicts))

        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = float(time_budget)
        # Use all cores for a large hall; ignore if the field name differs.
        try:
            solver.parameters.num_workers = 8
        except Exception:
            pass

        class _Canceller(cp_model.CpSolverSolutionCallback):
            def __init__(self, ev):
                super().__init__()
                self._ev = ev

            def on_solution_callback(self):
                if self._ev.is_set():
                    self.StopSearch()

        status = solver.Solve(model, _Canceller(cancel_event))

        if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            # No usable solution within budget → heuristic fallback.
            return self._greedy_fallback(m, n, students, neighbor_pairs, cancel_event, time_budget)

        # Determine which class sits in each seat.
        class_seats = {c: [] for c in class_names}
        for j in range(n):
            for c in class_names:
                if solver.Value(y[(j, c)]) == 1:
                    class_seats[c].append(j)
                    break

        # Map students → seats. Sort each class's seats by seat number so roll
        # numbers flow in a natural reading order across the hall.
        assignment = [0] * m
        for c in class_names:
            seats_c = sorted(class_seats[c], key=lambda j: seat_numbers[j])
            for idx, k in enumerate(classes[c]):
                assignment[k] = seats_c[idx]

        cost = int(round(solver.ObjectiveValue()))
        return assignment, cost

    # ───────────────────── heuristic fallback ─────────────────────
    def _greedy_fallback(self, m, n, students, neighbor_pairs, cancel_event, time_budget):
        """
        Targeted hill-climb, used only if CP-SAT is unavailable or errors out.
        Unlike the old purely-random version, it relocates students that are
        *actually* in conflict, so it converges far faster.
        """
        best = random.sample(range(n), m)
        best_cost = self._count_conflicts(best, students, neighbor_pairs)
        start = time.time()

        while time.time() - start < time_budget and best_cost > 0:
            if cancel_event.is_set():
                break
            occupant = {best[k]: k for k in range(m)}
            conflicted = []
            for (a, b) in neighbor_pairs:
                ka, kb = occupant.get(a), occupant.get(b)
                if ka is not None and kb is not None and students[ka]["Class"] == students[kb]["Class"]:
                    conflicted.append(ka)
                    conflicted.append(kb)
            if not conflicted:
                break
            ka = random.choice(conflicted)
            kb = random.randrange(m)
            if ka == kb:
                continue
            trial = best[:]
            trial[ka], trial[kb] = trial[kb], trial[ka]
            cost = self._count_conflicts(trial, students, neighbor_pairs)
            if cost < best_cost:
                best, best_cost = trial, cost

        return best, best_cost

    def _count_conflicts(self, assignment, students, neighbor_pairs):
        """Count # of same-class adjacent pairs given assignment[] and neighbor_pairs."""
        # build reverse map seat_idx->student_idx
        occupant = {assignment[k]: k for k in range(len(assignment))}
        conflicts = 0
        for i, j in neighbor_pairs:
            ki = occupant.get(i, None)
            kj = occupant.get(j, None)
            if ki is not None and kj is not None:
                if students[ki]["Class"] == students[kj]["Class"]:
                    conflicts += 1
        return conflicts

    def _on_assignment_done(self, assignment, students, fallback):
        # close progress
        # (we stored it in a local so just find any open QProgressDialog)
        for w in self.findChildren(QProgressDialog):
            w.close()

        # Get only unlocked and selected seats (same logic as in shuffle_and_assign_roll_numbers)
        available_seats = [cw for row in self.hall_layout for cw in row if cw.selected and not cw.locked]
        m = len(assignment)

        self.push_undo()
        for cw in available_seats:
            cw.clear()

        for k in range(m):
            idx = assignment[k]
            if idx < len(available_seats):
                cw = available_seats[idx]
                cw.student_data = dict(students[k])
                cw.update()

        self.detect_nearby()
        self.refresh_legend()

        if fallback:
            self.show_message_box(
                "جزوي بهینه کول",
                "بهینه کول بشپړ شول خو ځینې یو شان صنفونه نږدې پاتې شول.\n"
                "تاسو کولی شئ بیا هڅه وکړئ یا لاسي بدل کړئ.",
            )
        else:
            self.show_message_box(
                "بشپړ شو", "بریالیه د صفر یو شان صنف نږدې والي سره تخصیص شول!"
            )

    def show_class_color_menu(self, class_name, pos):
        """Show context menu for changing class color"""
        menu = QMenu(self)
        change_color_action = menu.addAction("رنګ بدل کړئ")
        change_color_action.triggered.connect(lambda: self.change_class_color(class_name))
        menu.exec_(pos)

    def change_class_color(self, class_name):
        """Change color for a specific class"""
        if class_name not in self.class_colors:
            return
            
        color = QColorDialog.getColor(QColor(f"#{self.class_colors[class_name]}"), self, f"د {class_name} لپاره رنګ انتخاب کړئ")
        if color.isValid():
            self.class_colors[class_name] = color.name()[1:]  # Remove # from hex
            self.refresh_legend()
            # Update all cells with this class
            for row in self.hall_layout:
                for cw in row:
                    if cw.student_data and cw.student_data["Class"] == class_name:
                        cw.update()

    # ───────────────────── Other features ─────────────────────
    def change_grid_size(self):
        r, ok = QInputDialog.getInt(self, "قطارونه", "د قطارونو شمیر:", self.rows, 1, 100)
        if not ok:
            return
        c, ok = QInputDialog.getInt(self, "ستونونه", "د ستونونو شمیر:", self.cols, 1, 100)
        if not ok:
            return
        self.push_undo()
        self.rows, self.cols = r, c
        self.rebuild_grid()

    def push_undo(self):
        import pickle

        state = pickle.dumps(
            {
                "rows": self.rows,
                "cols": self.cols,
                "class_colors": self.class_colors,
                "cells": [
                    {"r": r, "c": c, "stu": cw.student_data, "locked": cw.locked}
                    for r, row in enumerate(self.hall_layout)
                    for c, cw in enumerate(row)
                ],
            }
        )
        self.undo_stack.append(state)
        if len(self.undo_stack) > UNDO_LIMIT:
            self.undo_stack.pop(0)
        self.redo_stack.clear()

    def undo(self):
        if not self.undo_stack:
            return
        import pickle

        self.redo_stack.append(self.undo_stack.pop())
        state = pickle.loads(self.redo_stack[-1])
        self._restore_state(state)

    def redo(self):
        if not self.redo_stack:
            return
        import pickle

        self.undo_stack.append(self.redo_stack.pop())
        state = pickle.loads(self.undo_stack[-1])
        self._restore_state(state)

    def update_undo_buttons(self):
        pass # No longer needed

    def on_seat_change(self):
        self.push_undo()

    def _restore_state(self, d):
        self.rows, self.cols = d["rows"], d["cols"]
        self.class_colors = d["class_colors"]
        self.rebuild_grid()
        for rec in d["cells"]:
            cw = self.hall_layout[rec["r"]][rec["c"]]
            cw.student_data = rec["stu"]
            cw.locked = rec["locked"]
            cw.update()
        self.detect_nearby()
        self.refresh_legend()

    def set_class_colors(self):
        if not self.student_data:
            self.show_message_box("هیڅ معلومات نشته", "مهرباني وکړئ لومړی زده کوونکي وګرځوئ.", QMessageBox.Warning)
            return
        classes = {d["Class"] for d in self.student_data}
        for cls in classes:
            col = QColorDialog.getColor()
            if col.isValid():
                self.class_colors[cls] = col.name()[1:]
        self.refresh_legend()

    def detect_nearby(self):
        nearby = {}
        for r in range(self.rows):
            for c in range(self.cols):
                cw = self.hall_layout[r][c]
                if cw.student_data:
                    cls = cw.student_data["Class"]
                    for dr, dc in ADJ_DIRECTIONS:
                        nr, nc = r + dr, c + dc
                        if 0 <= nr < self.rows and 0 <= nc < self.cols:
                            cw2 = self.hall_layout[nr][nc]
                            if cw2.student_data and cw2.student_data["Class"] == cls:
                                nearby[cls] = nearby.get(cls, 0) + 1
        for cls in nearby:
            nearby[cls] //= 2
        self.nearby_counts = nearby

    def refresh_legend(self):
        self.detect_nearby()
        while self.legend_layout.count():
            w = self.legend_layout.takeAt(0).widget()
            if w:
                w.deleteLater()
        maxp, wrap = 8, 10
        for i, (cls, col) in enumerate(self.class_colors.items()):
            cnt = self.nearby_counts.get(cls, 0)
            txt = f"{cls} ({cnt} adj)" if cnt else cls
            btn = QPushButton("\n".join(textwrap.wrap(txt, wrap)))
            btn.setFixedWidth(100)
            btn.setFixedHeight(40)
            btn.setStyleSheet(f"background-color:#{col};border:1px solid black;padding:2px;font-size:10pt;")
            # Connect to both class info and color change
            btn.clicked.connect(lambda checked, class_name=cls: self.show_class_info(class_name))
            btn.setContextMenuPolicy(Qt.CustomContextMenu)
            btn.customContextMenuRequested.connect(lambda pos, class_name=cls: self.show_class_color_menu(class_name, pos))
            r, c = divmod(i, maxp)
            self.legend_layout.addWidget(btn, r, c)

    def show_class_info(self, cls):
        """
        Display a dialog listing every student in `cls` along with:
        • Seat number
        • Roll number
        • Student name
        • Father name
        • Adjacent seats (same class)
        """
        # 1) Gather data
        rows = []
        for r in range(self.rows):
            for c in range(self.cols):
                cw = self.hall_layout[r][c]
                if cw.student_data and cw.student_data["Class"] == cls:
                    seat = cw.number
                    d = cw.student_data
                    # find adjacent seats
                    adjs = []
                    for dr, dc in ADJ_DIRECTIONS:
                        nr, nc = r + dr, c + dc
                        if 0 <= nr < self.rows and 0 <= nc < self.cols:
                            cw2 = self.hall_layout[nr][nc]
                            if cw2.student_data and cw2.student_data["Class"] == cls:
                                adjs.append(str(cw2.number))
                    rows.append(
                        {
                            "Seat": seat,
                            "Roll": d.get("RollNumber", ""),
                            "Name": d.get("Name", ""),
                            "Father": d.get("Father", ""),
                            "Adjacent": ", ".join(adjs) or "None",
                        }
                    )

        # 2) Build dialog
        dlg = QDialog(self)
        dlg.setWindowTitle(f"{cls} — Student Details")
        dlg.resize(600, 400)

        tbl = QTableWidget(len(rows), 5, dlg)
        tbl.setHorizontalHeaderLabels(["Seat", "Roll Number", "Name", "Father", "Adjacent Seats"])
        tbl.verticalHeader().setVisible(False)
        tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        tbl.setSelectionBehavior(QTableWidget.SelectRows)
        tbl.setSelectionMode(QTableWidget.SingleSelection)
        tbl.setAlternatingRowColors(True)
        tbl.setWordWrap(True)
        tbl.setColumnWidth(0, 60)
        tbl.setColumnWidth(1, 100)
        tbl.setColumnWidth(4, 120)

        # 3) Populate
        for i, rec in enumerate(rows):
            tbl.setItem(i, 0, QTableWidgetItem(str(rec["Seat"])))
            tbl.setItem(i, 1, QTableWidgetItem(str(rec["Roll"])))
            tbl.setItem(i, 2, QTableWidgetItem(rec["Name"]))
            tbl.setItem(i, 3, QTableWidgetItem(rec["Father"]))
            tbl.setItem(i, 4, QTableWidgetItem(rec["Adjacent"]))

        tbl.resizeColumnsToContents()

        # 4) Layout + Close button
        btn_close = QPushButton("Close")
        btn_close.clicked.connect(dlg.close)

        vlay = QVBoxLayout(dlg)
        vlay.addWidget(tbl, 1)
        hlay = QHBoxLayout()
        hlay.addStretch()
        hlay.addWidget(btn_close)
        vlay.addLayout(hlay)

        dlg.exec_()

    def show_cell_menu(self, cell):
        menu = QMenu(self)

        # your existing actions
        if self.moving_cell and self.moving_cell != cell and not cell.locked:
            menu.addAction("Assign Here", lambda: self.try_swap(cell))
        if cell.student_data:
            menu.addAction("Cut Student", lambda: self.cut_student(cell))
            menu.addAction("Clear Seat", lambda: self.clear_seat(cell))
        if self.copied_data and not cell.student_data and not cell.locked:
            menu.addAction("Paste Here", lambda: self.paste_student(cell))
        menu.addAction(
            "Lock Seat" if not cell.locked else "Unlock Seat", lambda: cell.toggle_lock()
        )

        # ─── New “Deselect All” action ───────────────────────
        menu.addSeparator()
        menu.addAction("Deselect All", self._deselect_all)

        menu.exec_(QCursor.pos())

    def _deselect_all(self):
        """Clear the selected flag on every CellWidget."""
        for row in self.hall_layout:
            for cw in row:
                if cw.selected:
                    cw.selected = False
                    cw.update()

    def start_move(self, cell):
        self.moving_cell = cell
        cell.selected = True
        cell.update()

    def try_swap(self, tgt):
        if not self.moving_cell or tgt.locked:
            return
        self.push_undo()
        src = self.moving_cell
        src.student_data, tgt.student_data = tgt.student_data, src.student_data
        src.selected = False
        src.update()
        tgt.update()
        self.moving_cell = None

    def cut_student(self, cell):
        if not cell.student_data:
            return
        self.push_undo()
        self.copied_data = dict(cell.student_data)
        cell.clear()

    def paste_student(self, cell):
        if not self.copied_data or cell.student_data or cell.locked:
            return
        self.push_undo()
        cell.student_data = dict(self.copied_data)
        cell.update()

    def clear_seat(self, cell):
        self.push_undo()
        cell.clear()

    def open_select_classes_dialog(self):
        exam_id = self.cb_exam.currentData()
        if not exam_id:
            self.show_message_box("هیڅ امتحان نشته", "مهرباني وکړئ لومړی یو امتحان انتخاب کړئ.", QMessageBox.Warning)
            return

        # fetch all class-exam entries for this exam
        conn = DBManager.get_connection()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT ce.id, c.class_name
            FROM dbo.ClassExams ce
            JOIN dbo.classes    c ON ce.class_id = c.ID
            WHERE ce.exam_id = ?
        """,
            exam_id,
        )
        options = cur.fetchall()  # [(class_exam_id, class_name), …]
        conn.close()

        # build dialog with checkboxes
        dlg = QDialog(self)
        dlg.setWindowTitle("صنفونه انتخاب کړئ")
        layout = QVBoxLayout(dlg)
        checkboxes = []
        for class_exam_id, class_name in options:
            cb = QCheckBox(f"{class_name}  (ID: {class_exam_id})")
            cb.class_exam_id = class_exam_id
            cb.setFont(self.input_font)
            layout.addWidget(cb)
            checkboxes.append(cb)

        # OK / Cancel
        btn_box = QHBoxLayout()
        btn_ok = QPushButton("سمه ده")
        btn_ok.setFont(self.input_font)
        btn_ok.setStyleSheet(self.button_style)
        btn_cancel = QPushButton("لغوه کړئ")
        btn_cancel.setFont(self.input_font)
        btn_cancel.setStyleSheet(self.button_style)
        btn_box.addWidget(btn_ok)
        btn_box.addWidget(btn_cancel)
        layout.addLayout(btn_box)

        btn_ok.clicked.connect(dlg.accept)
        btn_cancel.clicked.connect(dlg.reject)

        if dlg.exec_() == QDialog.Accepted:
            # gather selections
            self.selected_class_exam_ids = [cb.class_exam_id for cb in checkboxes if cb.isChecked()]
            count = len(self.selected_class_exam_ids)
            if count == 0:
                self.show_message_box(
                    "هیڅ صنفونه انتخاب نه شول", "مهرباني وکړئ لومړی صنفونه انتخاب کړئ."
                )
                return

            # update button text & load students
            self.btn_select_classes.setText(f"{count} صنف(ونه) انتخاب شول")
            self.load_students_from_db()

    def load_students_from_db(self):
        """Query Admissions → general_registration → ClassExams → classes
        for all selected class_exam_ids, and load into self.student_data."""
        ids = self.selected_class_exam_ids
        placeholders = ",".join("?" for _ in ids)
        sql = f"""
            SELECT
            a.student_id,
            gr.name         AS student_name,
            gr.father_name  AS father_name,
            gr.card_number  AS card_number,
            c.class_name
            FROM dbo.Admissions a
            JOIN dbo.general_registration gr ON a.student_id = gr.ID
            JOIN dbo.ClassExams ce            ON a.class_exam_id = ce.id
            JOIN dbo.classes    c            ON ce.class_id       = c.ID
            WHERE ce.id IN ({placeholders})
            ORDER BY c.class_name, gr.card_number
        """
        conn = DBManager.get_connection()
        cur = conn.cursor()
        cur.execute(sql, *ids)
        rows = cur.fetchall()
        conn.close()

        # Map to the same dict‐shape your code expects:
        self.student_data = [
            {"ID": sid, "RollNumber": card, "Class": cls, "Name": name, "Father": father}
            for sid, name, father, card, cls in rows
        ]
        self.show_message_box(
            "زده کوونکي وګرځول شول",
            f"{len(self.student_data)} زده کوونکي له {len(ids)} صنف(ونو) څخه وګرځول شول.",
        )

    # ───────────────────── search ───────────────────────────────
    def perform_search(self):
        """Perform search when button is clicked or Enter is pressed"""
        search_text = self.search_box.text().strip()
        if search_text:
            self.search_student(search_text)
        else:
            self.clear_search_highlights()

    def clear_search_highlights(self):
        """Clear all search highlights"""
        for row in self.hall_layout:
            for cw in row:
                cw.search_highlight = False
                cw.update()

    def search_student(self, text):
        t = text.lower()
        found_students = []
        found_cells = []
        
        # Clear previous highlights first
        self.clear_search_highlights()
        
        for row in self.hall_layout:
            for cw in row:
                if cw.student_data and (
                    t in str(cw.student_data["ID"]).lower() or t in cw.student_data["Name"].lower()
                ):
                    found_students.append(cw.student_data)
                    found_cells.append(cw)
        
        # Highlight found cells with red border
        for cw in found_cells:
            col = "#" + self.get_class_color(cw.student_data["Class"])
            cw.search_highlight = True # Set the flag
            cw.update()
        
        # Show results
        if found_students:
            if len(found_students) == 1:
                student = found_students[0]
                student_info = f"ID: {student['ID']}\nنوم: {student['Name']}\nد پلار نوم: {student['Father']}\nصنف: {student['Class']}\nد رقم الجلوس: {student.get('RollNumber', 'N/A')}"
                self.search_box.setToolTip(student_info)
                self.show_message_box("موندل شو", f"زده کوونکی وموندل شو:\n{student_info}")
            else:
                student_info = f"ID: {found_students[0]['ID']}\nنوم: {found_students[0]['Name']}\nد پلار نوم: {found_students[0]['Father']}\nصنف: {found_students[0]['Class']}\nد رقم الجلوس: {found_students[0].get('RollNumber', 'N/A')}"
                self.search_box.setToolTip(student_info)
                self.show_message_box("موندل شول", f"{len(found_students)} زده کوونکي وموندل شول\nلومړی زده کوونکی:\n{student_info}")
        else:
            self.search_box.setToolTip("")
            self.show_message_box("موندل نه شو", "هیڅ زده کوونکی وموندل نه شو.")

    # ───────────────────── Excel export ─────────────────────────────────
    def generate_excel(self):
        """
        Exports two sheets:
        1) ‘Seating Map’ with the exact layout, colors, and student info.
        2) ‘Student List’ with one row per occupied seat.
        Title and font size adjust based on the selected exam and grid dimensions.
        """
        # 1) Make sure we have at least one color mapping (implies a map is loaded or created)
        if not self.class_colors:
            self.show_message_box("نیمګړی معلومات", "مهرباني وکړئ لومړی یو نقشه وګرځوئ او رنګونه تنظیم کړئ.", QMessageBox.Warning)
            return

        try:
            # 2) Prepare workbook & common styles
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Seating Map"
            align_all = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # 3) Dynamic title font size (base 45 at 20 columns)
            base_size = 45
            hdr_size = max(14, int(base_size * self.cols / 20))
            hdr_font = Font(name="Bahij Titr", size=hdr_size, bold=True)

            # 4) Dynamic title text using the exam name
            exam_name = self.cb_exam.currentText() or ""
            title_text = f"{exam_name} - د رقم الجلوس نقشه"

            # 5) Merge the top row and set the title
            last_col = get_column_letter(self.cols)
            ws.merge_cells(f"A1:{last_col}1")
            ws["A1"].value = title_text
            ws["A1"].font = hdr_font
            ws["A1"].alignment = align_all

            # 6) Fill the seating‐map grid
            for r in range(self.rows):
                for c in range(self.cols):
                    cw = self.hall_layout[r][c]
                    seat_no = cw.number

                    if cw.student_data:
                        d = cw.student_data
                        info = "\n".join(
                            [
                                str(seat_no),
                                str(d["ID"]),
                                str(d["Name"]),
                                str(d["Father"]),
                                str(d["Class"]),
                            ]
                        )
                        fill_hex = self.class_colors.get(d["Class"], "FFFFFF")
                    else:
                        info = str(seat_no)
                        fill_hex = "FFFFFF"

                    cell = ws.cell(row=r + 2, column=c + 1, value=info)
                    cell.font = Font(name="Bahij Titr", size=14)
                    cell.alignment = align_all
                    cell.border = border
                    cell.fill = PatternFill(
                        start_color=fill_hex, end_color=fill_hex, fill_type="solid"
                    )

            # 7) Adjust row heights & column widths
            ws.row_dimensions[1].height = hdr_size * 1.5  # scale height to match font
            for rr in range(2, self.rows + 2):
                ws.row_dimensions[rr].height = 80
            for cc in range(1, self.cols + 1):
                ws.column_dimensions[get_column_letter(cc)].width = 15

            # 8) Page setup for A3
            ws.page_setup.paperSize = ws.PAPERSIZE_A3
            ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)

            # 9) Create the ‘Student List’ sheet
            sl = wb.create_sheet(title="Student List")
            sl.merge_cells("A1:E1")
            sl["A1"].value = "د شاګردانو معلومات"
            sl["A1"].font = Font(name="Bahij Nassim", size=14, bold=True)
            sl["A1"].alignment = align_all

            # 10) Headers, with “Seat” → “رقم الجلوس”
            headers = ["ID", "رقم الجلوس", "نوم", "د پلار نوم", "درجه"]
            for idx, hdr in enumerate(headers, start=1):
                cell = sl.cell(row=2, column=idx, value=hdr)
                cell.font = Font(name="Bahij Nassim", size=14, bold=True)
                cell.alignment = align_all
                cell.border = border

            # 11) Gather & sort occupied seats
            occupants = []
            for r in range(self.rows):
                for c in range(self.cols):
                    cw = self.hall_layout[r][c]
                    if cw.student_data:
                        d = cw.student_data
                        occupants.append(
                            {
                                "ID": d["ID"],
                                "Seat": cw.number,
                                "Name": d["Name"],
                                "Father": d["Father"],
                                "Class": d["Class"],
                            }
                        )
            occupants.sort(key=lambda x: x["Seat"])

            # 12) Write them out
            row_cursor = 3
            for oc in occupants:
                sl.cell(row=row_cursor, column=1, value=oc["ID"])
                sl.cell(row=row_cursor, column=2, value=oc["Seat"])
                sl.cell(row=row_cursor, column=3, value=oc["Name"])
                sl.cell(row=row_cursor, column=4, value=oc["Father"])
                sl.cell(row=row_cursor, column=5, value=oc["Class"])
                for col_idx in range(1, 6):
                    cell = sl.cell(row=row_cursor, column=col_idx)
                    cell.font = Font(name="Bahij Nassim", size=14)
                    cell.alignment = align_all
                    cell.border = border
                row_cursor += 1

            # 13) Final layout for the list sheet
            for cc in range(1, 6):
                sl.column_dimensions[get_column_letter(cc)].width = 15
            sl.page_setup.paperSize = sl.PAPERSIZE_A4
            sl.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)

            # 14) Prompt save
            path, _ = QFileDialog.getSaveFileName(self, "Excel خوندي کړئ", "", "Excel Files (*.xlsx)")
            if path:
                wb.save(path)
                self.show_message_box("بشپړ شو", "Excel فایل په بریالیتوب سره جوړ شو!")

        except Exception as e:
            self.show_message_box("تیروتنه", str(e), QMessageBox.Critical)

    def generate_html_report(self):
        # 1) Ensure we have a map
        if not self.class_colors:
            self.show_message_box("هیڅ معلومات نشته", "مهرباني وکړئ لومړی یو نقشه وګرځوئ یا جوړ کړئ.", QMessageBox.Warning)
            return

        # 2) Title & dynamic CSS with centered text
        exam_name = self.cb_exam.currentText() or "—"
        title_html = f"<h1>{exam_name} - د رقم الجلوس نقشه</h1>"

        # Choose a min‐cell size so things don’t get too cramped
        if self.cols <= 10:
            cell_px = 120
        elif self.cols <= 20:
            cell_px = 80
        else:
            cell_px = 60

        css = f"""
        <style>
        @page {{ size: auto; margin: 0; }}
        body {{ margin:0; font-family:'Bahij Titr',sans-serif; font-size:12pt; }}
        h1 {{ text-align:center; margin:16px 0; }}
        table.grid {{
            border-collapse: collapse;
            margin: auto;
            table-layout: auto;
            width: auto;
        }}
        table.grid td {{
            border: 1px solid #666;
            padding: 8px;
            vertical-align: middle;
            text-align: center;          /* <-- center text horizontally */
            word-break: break-word;
            line-height: 1.2;
            font-size:10pt;
            min-width: {cell_px}px;
            min-height: {cell_px}px;
        }}
        </style>
        """

        # 3) Build each row (including Class under Father)
        rows = []
        for r in range(self.rows):
            cells = []
            for c in range(self.cols):
                cw = self.hall_layout[r][c]
                if cw.student_data:
                    d = cw.student_data
                    bg = "#" + self.class_colors.get(d["Class"], "FFFFFF")
                    txt = (
                        f"{cw.number}<br/>"
                        f"ID: {d['ID']}<br/>"
                        f"{d['Name']}<br/>"
                        f"{d['Father']}<br/>"
                        f"<b>{d['Class']}</b>"
                    )
                else:
                    bg = "#AAAAAA" if cw.locked else "#ECF0F1"
                    txt = str(cw.number)
                cells.append(f"<td style='background:{bg};'>{txt}</td>")
            rows.append("<tr>" + "".join(cells) + "</tr>")

        table_html = "<table class='grid'>" + "".join(rows) + "</table>"

        # 4) Assemble the HTML
        html = (
            "<!DOCTYPE html><html dir='rtl'><head><meta charset='utf-8'>"
            + css
            + "</head><body>"
            + title_html
            + table_html
            + "</body></html>"
        )

        # 5) Show in dialog
        dlg = SeatingMapReportDialog(html, parent=self)
        dlg.exec_()

    # ───────────────────── DB save/load ──────────────────────────────────────
    def on_save_map(self):
        conn = DBManager.get_connection()
        cur = conn.cursor()

        # ─── New map vs Update existing ────────────────────────────
        if not self.current_map_id:
            # 1) Ask for a map name
            name, ok = QInputDialog.getText(self, "د نقشې نوم", "د نقشې نوم داخل کړئ:")
            if not ok or not name.strip():
                return

            # 2) Grab the selected exam ID
            eid = self.cb_exam.currentData()
            if eid is None:
                self.show_message_box("هیڅ امتحان نشته", "مهرباني وکړئ د خوندي کولو دمخه یو امتحان انتخاب کړئ.", QMessageBox.Warning)
                return

            # 3) Insert and get the new PK using SCOPE_IDENTITY()
            cur.execute(
                """
                INSERT INTO dbo.SeatingMaps (ExamID, MapName, Rows, Cols)
                VALUES (?, ?, ?, ?);
                SELECT SCOPE_IDENTITY() AS SeatingMapID;
            """,
                eid,
                name.strip(),
                self.rows,
                self.cols,
            )
            row = cur.fetchone()
            if row and row[0] is not None:
                self.current_map_id = int(row[0])
            else:
                self.show_message_box("تیروتنه", "نقشه خوندي کول ناکام شول.", QMessageBox.Critical)
                conn.rollback()
                conn.close()
                return

        else:
            # 4) Update dimensions on an existing map
            cur.execute(
                "UPDATE dbo.SeatingMaps SET Rows = ?, Cols = ? WHERE SeatingMapID = ?",
                self.rows,
                self.cols,
                self.current_map_id,
            )
            conn.commit()

            # 5) Wipe out old colors & assignments
            cur.execute(
                "DELETE FROM dbo.ClassColors     WHERE SeatingMapID = ?", self.current_map_id
            )
            cur.execute(
                "DELETE FROM dbo.SeatAssignments WHERE SeatingMapID = ?", self.current_map_id
            )

        # ─── Persist class colors ──────────────────────────────────
        for cls, hexcol in self.class_colors.items():
            cur.execute(
                "INSERT INTO dbo.ClassColors (SeatingMapID, ClassName, ColorHex) VALUES (?, ?, ?)",
                self.current_map_id,
                cls,
                hexcol,
            )

        # ─── Persist each seat assignment ─────────────────────────
        for r, row in enumerate(self.hall_layout):
            for c, cw in enumerate(row):
                sid = cw.student_data["ID"] if cw.student_data else None
                lk = 1 if cw.locked else 0
                cur.execute(
                    """
                    INSERT INTO dbo.SeatAssignments
                    (SeatingMapID, SeatRow, SeatCol, SeatNumber, Locked, StudentID)
                    VALUES (?, ?, ?, ?, ?, ?)
                """,
                    self.current_map_id,
                    r,
                    c,
                    cw.number,
                    lk,
                    sid,
                )

        # ─── Finalize ─────────────────────────────────────────────
        conn.commit()
        conn.close()
        self.show_message_box("خوندي شو", "نقشه په ډیټابیس کې خوندي شوه.")

    def new_map(self):
        """
        Clears current map state: selections, colors, loaded students, undo/redo, current_map_id,
        so you can start a brand-new seating map.
        """
        # 1) Reset map identity & undo history
        self.current_map_id = None
        self.undo_stack.clear()
        self.redo_stack.clear()

        # 2) Clear exam & class selections
        self.cb_exam.setCurrentIndex(0)
        self.selected_class_exam_ids = []
        self.btn_select_classes.setText("صنفونه انتخاب کړئ")
        self.student_data = []

        # 3) Clear class-color map & legend
        self.class_colors.clear()
        self.refresh_legend()

        # 4) Clear the grid: seats, locks, selections
        for row in self.hall_layout:
            for cw in row:
                cw.student_data = None
                cw.locked = False
                cw.selected = False
                cw.update()

        # 5) (Optional) Reset grid size back to default
        #    Uncomment if you want to also restore the default 40×20 size:
        # self.rows, self.cols = 40, 20
        # self.rebuild_grid()

        self.show_message_box("نوی نقشه", "د یو نوي د څوکیو نقشې لپاره چمتو دی.")

    def open_manage_maps(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("د خوندي شویو نقشو مدیریت")
        ly = QVBoxLayout(dlg)
        tbl = QTableWidget()
        tbl.setFont(self.input_font)
        ly.addWidget(tbl)
        conn = DBManager.get_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT SeatingMapID,ExamID,MapName,Rows,Cols,CreatedAt FROM dbo.SeatingMaps ORDER BY CreatedAt DESC"
        )
        recs = cur.fetchall()
        conn.close()
        tbl.setColumnCount(6)
        tbl.setHorizontalHeaderLabels(["ID", "د امتحان ID", "نوم", "قطارونه", "ستونونه", "د جوړیدو نیټه"])
        tbl.setRowCount(len(recs))
        for i, row in enumerate(recs):
            for j, v in enumerate(row):
                tbl.setItem(i, j, QTableWidgetItem(str(v)))
        tbl.resizeColumnsToContents()
        b1 = QPushButton("وګرځوئ")
        b1.setFont(self.input_font)
        b1.setStyleSheet(self.button_style)
        b2 = QPushButton("ړنګ کړئ")
        b2.setFont(self.input_font)
        b2.setStyleSheet(self.button_style)
        b3 = QPushButton("تړل")
        b3.setFont(self.input_font)
        b3.setStyleSheet(self.button_style)
        hl = QHBoxLayout()
        hl.addWidget(b1)
        hl.addWidget(b2)
        hl.addStretch()
        hl.addWidget(b3)
        ly.addLayout(hl)
        b1.clicked.connect(lambda: self._load_from_dialog(tbl, dlg))
        b2.clicked.connect(lambda: self._delete_from_dialog(tbl))
        b3.clicked.connect(dlg.close)
        dlg.exec_()

    def _load_from_dialog(self, tbl, dlg):
        r = tbl.currentRow()
        if r < 0:
            self.show_message_box("انتخاب", "هیڅ نقشه انتخاب نه شوه.", QMessageBox.Warning)
            return
        mid = int(tbl.item(r, 0).text())
        dlg.close()
        self.load_map(mid)

    def _delete_from_dialog(self, tbl):
        r = tbl.currentRow()
        if r < 0:
            self.show_message_box("انتخاب", "هیڅ نقشه انتخاب نه شوه.", QMessageBox.Warning)
            return
        mid = int(tbl.item(r, 0).text())
        reply = self.show_message_box(
            "تایید", f"نقشه {mid} ړنګ کړئ؟ ", QMessageBox.Question
        )
        if reply != QMessageBox.Yes:
            return
        conn = DBManager.get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM dbo.ClassColors    WHERE SeatingMapID=?", mid)
        cur.execute("DELETE FROM dbo.SeatAssignmentsWHERE SeatingMapID=?", mid)
        cur.execute("DELETE FROM dbo.SeatingMaps     WHERE SeatingMapID=?", mid)
        conn.commit()
        conn.close()
        tbl.removeRow(r)
        self.show_message_box("ړنګ شو", "نقشه ړنګه شوه.")

    def load_map(self, map_id: int):
        conn = DBManager.get_connection()
        cur = conn.cursor()

        # 1) Fetch saved map metadata
        cur.execute(
            """
            SELECT ExamID, MapName, Rows, Cols
            FROM dbo.SeatingMaps
            WHERE SeatingMapID = ?
        """,
            map_id,
        )
        row = cur.fetchone()
        if not row:
            self.show_message_box("تیروتنه", "نقشه و نه موندل شوه.", QMessageBox.Critical)
            conn.close()
            return

        eid, map_name, rows, cols = row
        self.current_map_id = map_id
        self.rows, self.cols = rows, cols

        # 1.1) Update exam dropdown to match this map’s exam
        idx = self.cb_exam.findData(eid)
        if idx != -1:
            # block the signal so we don’t clear out student_data inadvertently
            self.cb_exam.blockSignals(True)
            self.cb_exam.setCurrentIndex(idx)
            self.cb_exam.blockSignals(False)

        # 2) Rebuild the grid to the saved dimensions
        self.rebuild_grid()

        # 3) Load class‐to‐color mappings
        cur.execute(
            """
            SELECT ClassName, ColorHex
            FROM dbo.ClassColors
            WHERE SeatingMapID = ?
        """,
            map_id,
        )
        self.class_colors = {r[0]: r[1] for r in cur.fetchall()}

        # 4) Load each seat assignment record
        cur.execute(
            """
            SELECT SeatRow, SeatCol, Locked, StudentID
            FROM dbo.SeatAssignments
            WHERE SeatingMapID = ?
        """,
            map_id,
        )
        assignments = cur.fetchall()

        for seat_row, seat_col, locked_flag, student_id in assignments:
            try:
                cw = self.hall_layout[seat_row][seat_col]
            except IndexError:
                continue
            cw.locked = bool(locked_flag)

            if student_id is not None:
                c2 = conn.cursor()
                c2.execute(
                    f"""
                    SELECT student_id,
                        card_number,
                        class_name,
                        student_name,
                        father_name
                    FROM {VIEW_STUDS}
                    WHERE exam_id = ? AND student_id = ?
                """,
                    eid,
                    student_id,
                )
                rec = c2.fetchone()
                if rec:
                    sid, card_no, cls_name, stud_name, father = rec
                    cw.student_data = {
                        "ID": sid,
                        "RollNumber": card_no,
                        "Class": cls_name,
                        "Name": stud_name,
                        "Father": father,
                    }
            cw.update()

        conn.close()

        # 5) Finally, recalc adjacency & refresh legend colors
        self.detect_nearby()
        self.refresh_legend()

    def show_message_box(self, title, message, icon=QMessageBox.Information):
        """Helper method for consistent message boxes (EXACTLY as in general registration)"""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        msg_box.setIcon(icon)
        msg_box.setStyleSheet(self.message_box_style)
        
        # Add Yes/No buttons for question dialogs
        if icon == QMessageBox.Question:
            msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msg_box.setDefaultButton(QMessageBox.No)
        
        return msg_box.exec_()

    def get_class_color(self, cls):
        return self.class_colors.get(cls, "FFFFFF")

    def deselect_all(self):
        """Clear the selected flag on every CellWidget."""
        for row in self.hall_layout:
            for cw in row:
                if cw.selected:
                    cw.selected = False
                    cw.update()

    def _mk_action(self, text, slot):
        """Helper method to create toolbar actions"""
        action = QAction(text, self)
        action.triggered.connect(slot)
        return action

    def _mk_btn(self, text, slot):
        """Helper method to create consistent buttons"""
        btn = QPushButton(text)
        btn.setFont(self.input_font)
        btn.setStyleSheet(self.button_style)
        btn.clicked.connect(slot)
        return btn


def main():
    app = QApplication(sys.argv)
    try:
        from nazim.core.qt_startup import prime_desktop
        prime_desktop()
    except Exception:
        pass
    w = RollNumberMappingApp()
    w.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
